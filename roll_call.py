import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import random
import threading
import time
from typing import List, Dict

class Student:
    def __init__(self, name: str, gender: str, seat: int):
        self.name = name
        self.gender = gender
        self.seat = seat

class RollCallApp:
    def __init__(self, root):
        self.root = root
        self.root.title("课堂点名系统")
        self.root.geometry("650x600")
        self.root.resizable(False, False)
        
        # 数据存储
        self.students: List[Student] = []
        self.is_running = False # 动画是否在运行
        
        # 界面布局
        self.create_widgets()
        
    def create_widgets(self):
        # --- 1. 顶部：文件操作区 ---
        file_frame = ttk.LabelFrame(self.root, text="📂 学生信息管理", padding=15)
        file_frame.pack(fill="x", padx=20, pady=10)
        
        btn_style = ttk.Style()
        btn_style.configure('Big.TButton', font=('Arial', 10))
        
        ttk.Button(file_frame, text="生成模板", command=self.create_template, style='Big.TButton').pack(side="left", padx=5)
        ttk.Button(file_frame, text="导入名单", command=self.load_students, style='Big.TButton').pack(side="left", padx=5)
        self.file_label = ttk.Label(file_frame, text="状态：未导入文件", foreground="gray")
        self.file_label.pack(side="left", padx=15)
        
        # --- 2. 中部：筛选条件区 ---
        setting_frame = ttk.LabelFrame(self.root, text="⚙️ 抽取设置", padding=15)
        setting_frame.pack(fill="x", padx=20, pady=5)
        
        # 第一行：人数 & 性别
        row1 = ttk.Frame(setting_frame)
        row1.pack(fill="x", pady=5)
        
        ttk.Label(row1, text="抽取人数：").pack(side="left")
        self.count_var = tk.StringVar(value="1")
        ttk.Entry(row1, textvariable=self.count_var, width=5).pack(side="left", padx=5)
        
        ttk.Label(row1, text="性别：").pack(side="left", padx=(10, 0))
        self.gender_var = tk.StringVar(value="全部")
        ttk.Combobox(row1, textvariable=self.gender_var, values=["全部", "男", "女"], state="readonly", width=6).pack(side="left", padx=5)
        
        # 第二行：座号
        row2 = ttk.Frame(setting_frame)
        row2.pack(fill="x", pady=5)
        
        ttk.Label(row2, text="座号模式：").pack(side="left")
        self.seat_type_var = tk.StringVar(value="全部")
        ttk.Combobox(row2, textvariable=self.seat_type_var, values=["全部", "单号", "双号"], state="readonly", width=6).pack(side="left", padx=5)
        
        ttk.Label(row2, text="范围：").pack(side="left", padx=(10, 0))
        self.start_seat_var = tk.StringVar(value="1")
        self.end_seat_var = tk.StringVar(value="99")
        ttk.Entry(row2, textvariable=self.start_seat_var, width=5).pack(side="left")
        ttk.Label(row2, text="-").pack(side="left")
        ttk.Entry(row2, textvariable=self.end_seat_var, width=5).pack(side="left")

        # --- 3. 中下部：抽奖展示区 ---
        display_frame = tk.Frame(self.root, bg="#f0f0f0", padx=20, pady=20)
        display_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # 大标题
        self.title_label = tk.Label(display_frame, text="准备就绪", font=("微软雅黑", 20, "bold"), bg="#f0f0f0")
        self.title_label.pack(pady=(0, 10))
        
        # 滚动名字的展示框
        self.show_frame = tk.Frame(display_frame, bg="white", relief="sunken", bd=2, height=150)
        self.show_frame.pack(fill="x")
        self.show_frame.pack_propagate(False) # 强制大小
        
        self.name_label = tk.Label(self.show_frame, text="???", font=("微软雅黑", 40, "bold"), bg="white", fg="#333")
        self.name_label.pack(expand=True)
        
        # --- 4. 底部：操作按钮 ---
        action_frame = tk.Frame(self.root, pady=15)
        action_frame.pack(fill="x")
        
        self.action_btn = tk.Button(action_frame, text="🚀 开始点名", command=self.toggle_roll_call, 
                                     bg="#2196F3", fg="white", font=("微软雅黑", 16, "bold"),
                                     width=15, height=2, relief="flat", cursor="hand2")
        self.action_btn.pack()

    def create_template(self):
        """生成模板"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "学生信息"
        ws.append(["姓名", "性别", "座号"])
        
        # 生成一些示例数据方便测试
        surnames = ["张", "李", "王", "刘", "陈", "杨", "黄", "赵", "周", "吴"]
        names = ["伟", "芳", "娜", "秀英", "敏", "静", "丽", "强", "磊", "洋"]
        
        for i in range(1, 51):
            name = random.choice(surnames) + random.choice(names)
            gender = random.choice(["男", "女"])
            ws.append([name, gender, i])
            
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="学生名单.xlsx")
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("成功", f"模板已生成，请填写后导入！\n已为你预填了50个示例数据。")

    def load_students(self):
        """导入数据"""
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not file_path: return
            
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            self.students = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] and row[2]:
                    try:
                        self.students.append(Student(str(row[0]), str(row[1]), int(row[2])))
                    except: pass
            
            self.file_label.config(text=f"✅ 已导入 {len(self.students)} 人", foreground="green")
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def get_filtered_students(self) -> List[Student]:
        """根据条件筛选学生"""
        filtered = []
        try:
            start = int(self.start_seat_var.get())
            end = int(self.end_seat_var.get())
        except:
            start, end = 1, 999

        for s in self.students:
            # 座号范围
            if not (start <= s.seat <= end): continue
            # 性别
            if self.gender_var.get() != "全部" and s.gender != self.gender_var.get(): continue
            # 单双号
            if self.seat_type_var.get() == "单号" and s.seat % 2 == 0: continue
            if self.seat_type_var.get() == "双号" and s.seat % 2 != 0: continue
            
            filtered.append(s)
        return filtered

    def toggle_roll_call(self):
        """开关按钮逻辑"""
        if not self.students:
            messagebox.showwarning("提示", "请先导入学生名单！")
            return

        if not self.is_running:
            # 开始
            self.is_running = True
            self.action_btn.config(text="🛑 停！", bg="#f44336") # 变红
            # 开启线程运行动画
            threading.Thread(target=self.run_lottery, daemon=True).start()
        else:
            # 停止
            self.is_running = False

    def run_lottery(self):
        """动画线程"""
        candidates = self.get_filtered_students()
        
        if not candidates:
            self.root.after(0, lambda: messagebox.showwarning("错误", "没有符合条件的学生！"))
            self.is_running = False
            self.root.after(0, self.reset_btn)
            return

        # 1. 快速滚动阶段
        speed = 0.05 
        while self.is_running:
            # 随机选一个名字显示
            pick = random.choice(candidates)
            # 更新UI需要在主线程，这里用after
            self.root.after(0, lambda p=pick: self.name_label.config(text=p.name))
            time.sleep(speed)
        
        # 2. 停止阶段，选定最终结果
        try:
            count = int(self.count_var.get())
        except:
            count = 1
            
        if count > len(candidates): count = len(candidates)
        
        final_winners = random.sample(candidates, count)
        
        # 显示结果
        self.root.after(0, lambda: self.show_result(final_winners))
        self.root.after(0, self.reset_btn)

    def reset_btn(self):
        self.action_btn.config(text="🚀 开始点名", bg="#2196F3")

    def show_result(self, winners):
        self.title_label.config(text="🎉 中奖名单 🎉", fg="#d32f2f")
        
        if len(winners) == 1:
            s = winners[0]
            self.name_label.config(text=f"{s.name}", fg="#E91E63")
        else:
            # 多人中奖，拼接字符串
            text = "\n".join([f"{s.name} ({s.seat}号)" for s in winners])
            self.name_label.config(text=text, font=("微软雅黑", 20, "bold"), fg="#333")

if __name__ == "__main__":
    root = tk.Tk()
    app = RollCallApp(root)
    root.mainloop()
